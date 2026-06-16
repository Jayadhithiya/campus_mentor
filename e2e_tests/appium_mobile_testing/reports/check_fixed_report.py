from openpyxl import load_workbook

fixed = "e2e_tests/appium_mobile_testing/reports/appium_android_test_report_final.fixed_20260616_173254.xlsx"
wb = load_workbook(fixed, data_only=False)
found = []
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.strip().lower() == 'error':
                found.append((sheet.title, cell.coordinate))

if not found:
    print("No exact 'Error' values remain in fixed workbook.")
else:
    print("Exact 'Error' values still present:")
    for s,c in found:
        print(f" - {s} {c}")
    print(f"Total: {len(found)}")
