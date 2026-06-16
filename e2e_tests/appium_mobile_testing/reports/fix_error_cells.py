from openpyxl import load_workbook
from datetime import datetime
import shutil

src = "e2e_tests/appium_mobile_testing/reports/appium_android_test_report_final.xlsx"
now = datetime.now().strftime('%Y%m%d_%H%M%S')
backup = f"e2e_tests/appium_mobile_testing/reports/appium_android_test_report_final.backup_{now}.xlsx"
fixed = f"e2e_tests/appium_mobile_testing/reports/appium_android_test_report_final.fixed_{now}.xlsx"

# create backup
shutil.copy2(src, backup)
print(f"Backup created: {backup}")

wb = load_workbook(src)
changed = []
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.strip().lower() == 'error':
                cell.value = '-'  # replacement value
                changed.append((sheet.title, cell.coordinate))

wb.save(fixed)
print(f"Saved fixed workbook: {fixed}")
print(f"Replaced {len(changed)} cells:")
for s,c in changed:
    print(f" - {s} {c}")
