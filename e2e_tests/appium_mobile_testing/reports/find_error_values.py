from openpyxl import load_workbook
import re

wb_path = "e2e_tests/appium_mobile_testing/reports/appium_android_test_report_final.xlsx"
wb = load_workbook(wb_path, data_only=False)

errors = []
err_pattern = re.compile(r"#\w+|error", re.IGNORECASE)
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and err_pattern.search(val):
                errors.append((sheet.title, cell.coordinate, val))

if not errors:
    print("No obvious error strings found.")
else:
    for s, coord, val in errors:
        print(f"Sheet: {s} Cell: {coord} Value: {val}")
    print(f"Found {len(errors)} cells with error-like values.")
