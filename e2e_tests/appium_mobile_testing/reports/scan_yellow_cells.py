from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb_path = "e2e_tests/appium_mobile_testing/reports/appium_android_test_report_final.xlsx"
wb = load_workbook(wb_path, data_only=False)

results = []
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            fill = cell.fill
            # Detect any non-default fill
            try:
                patt = fill.patternType
            except Exception:
                patt = None
            fg = None
            try:
                fg = fill.fgColor.rgb or fill.fgColor.indexed or fill.fgColor.theme
            except Exception:
                fg = None
            if patt or fg:
                results.append((sheet.title, cell.coordinate, cell.value, cell.data_type, patt, fg))

if not results:
    print("No non-default fills found.")
else:
    for s, coord, val, dtype, patt, fg in results:
        print(f"Sheet: {s} Cell: {coord} Pattern: {patt} Fill: {fg} Type: {dtype} Value: {val}")
    print(f"Found {len(results)} cells with non-default fills.")
