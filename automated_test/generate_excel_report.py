"""
generate_excel_report.py — Generate the 5-tab Excel evidence report.

Output: automated_test/security_test_report.xlsx

Tabs:
  1. Executive Summary
  2. Endpoint Inventory
  3. Findings
  4. Remediation Status
  5. Raw Results

Conditional formatting:
  Critical = Red      High = Orange    Medium = Yellow
  Low = Blue          PASS  = Green    FAIL = Red fill
"""

import os
import json
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill,
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.dirname(os.path.abspath(__file__))
REPORT_PATH = os.path.join(OUT_DIR, 'security_test_report.xlsx')

# ─────────────────────────────────────────────────────────────
# Colour palette
# ─────────────────────────────────────────────────────────────
C = {
    'critical_bg':  'FFC7CE',   # Red
    'critical_fg':  '9C0006',
    'high_bg':      'FFCC99',   # Orange
    'high_fg':      'BE4B00',
    'medium_bg':    'FFEB9C',   # Yellow
    'medium_fg':    '9C6500',
    'low_bg':       'BDD7EE',   # Blue
    'low_fg':       '1F4E79',
    'pass_bg':      'C6EFCE',   # Green
    'pass_fg':      '276221',
    'fail_bg':      'FFC7CE',
    'fail_fg':      '9C0006',
    'header_bg':    '1F2D3D',   # Dark navy
    'header_fg':    'FFFFFF',
    'tab1_accent':  '2E75B6',
    'alt_row':      'F2F6FC',
    'border':       'B8CCE4',
}


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, color='000000', size=11):
    return Font(bold=bold, color=color, size=size, name='Calibri')


def _border():
    thin = Side(style='thin', color=C['border'])
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_row(ws, headers, row=1, widths=None):
    """Write a styled header row."""
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = _fill(C['header_bg'])
        cell.font = _font(bold=True, color=C['header_fg'], size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _border()
        if widths:
            ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]
    ws.row_dimensions[row].height = 30


def _data_row(ws, values, row, alt=False):
    """Write a data row with alternating colour."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=str(val) if val is not None else '')
        if alt:
            cell.fill = _fill(C['alt_row'])
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = _border()
        cell.font = _font(size=10)


def _severity_fill(sev):
    sev = str(sev).strip().lower()
    mapping = {
        'critical': (C['critical_bg'], C['critical_fg']),
        'high':     (C['high_bg'],     C['high_fg']),
        'medium':   (C['medium_bg'],   C['medium_fg']),
        'low':      (C['low_bg'],      C['low_fg']),
        'pass':     (C['pass_bg'],     C['pass_fg']),
        'fail':     (C['fail_bg'],     C['fail_fg']),
    }
    return mapping.get(sev, (None, None))


# ─────────────────────────────────────────────────────────────
# Tab 1 — Executive Summary
# ─────────────────────────────────────────────────────────────

def _tab_executive_summary(wb, test_results, endpoints):
    ws = wb.create_sheet('Executive Summary')
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells('A1:B1')
    title = ws['A1']
    title.value = '🛡  Security Test Report — Executive Summary'
    title.font = Font(bold=True, size=16, color=C['header_fg'], name='Calibri')
    title.fill = _fill(C['header_bg'])
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:B2')
    ts_cell = ws['A2']
    ts_cell.value = f'Generated: {datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")}'
    ts_cell.font = _font(size=10, color='666666')
    ts_cell.alignment = Alignment(horizontal='center')

    # Counts
    total = len(test_results)
    passed = sum(1 for r in test_results if r['status'] == 'PASS')
    failed = sum(1 for r in test_results if r['status'] == 'FAIL')
    critical = sum(1 for r in test_results if r['severity'] == 'Critical')
    high = sum(1 for r in test_results if r['severity'] == 'High')
    medium = sum(1 for r in test_results if r['severity'] == 'Medium')
    low = sum(1 for r in test_results if r['severity'] == 'Low')

    metrics = [
        ('Endpoints Tested',  len(endpoints)),
        ('Total Tests',        total),
        ('Passed ✅',          passed),
        ('Failed ❌',          failed),
        ('Critical 🔴',       critical),
        ('High 🟠',            high),
        ('Medium 🟡',          medium),
        ('Low 🔵',             low),
    ]

    _header_row(ws, ['Metric', 'Value'], row=4, widths=[30, 20])

    for i, (metric, value) in enumerate(metrics, 5):
        alt = (i % 2 == 0)
        _data_row(ws, [metric, value], row=i, alt=alt)
        # Colour-code value cell for severity counts
        val_cell = ws.cell(row=i, column=2)
        label_lower = metric.lower()
        if 'critical' in label_lower and value > 0:
            val_cell.fill = _fill(C['critical_bg'])
            val_cell.font = _font(bold=True, color=C['critical_fg'])
        elif 'high' in label_lower and value > 0:
            val_cell.fill = _fill(C['high_bg'])
            val_cell.font = _font(bold=True, color=C['high_fg'])
        elif 'medium' in label_lower and value > 0:
            val_cell.fill = _fill(C['medium_bg'])
            val_cell.font = _font(bold=True, color=C['medium_fg'])
        elif 'low' in label_lower and value > 0:
            val_cell.fill = _fill(C['low_bg'])
            val_cell.font = _font(bold=True, color=C['low_fg'])
        elif 'passed' in label_lower:
            val_cell.fill = _fill(C['pass_bg'])
            val_cell.font = _font(bold=True, color=C['pass_fg'])
        elif 'failed' in label_lower and value > 0:
            val_cell.fill = _fill(C['fail_bg'])
            val_cell.font = _font(bold=True, color=C['fail_fg'])
        val_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    return ws


# ─────────────────────────────────────────────────────────────
# Tab 2 — Endpoint Inventory
# ─────────────────────────────────────────────────────────────

def _tab_endpoint_inventory(wb, endpoints):
    ws = wb.create_sheet('Endpoint Inventory')
    ws.sheet_view.showGridLines = False

    headers = ['Method', 'Endpoint', 'Access Type', 'Expected Roles', 'Source File']
    widths = [14, 50, 22, 28, 40]
    _header_row(ws, headers, row=1, widths=widths)

    for i, ep in enumerate(endpoints, 2):
        alt = (i % 2 == 0)
        _data_row(ws, [
            ep.get('method', ''),
            ep.get('endpoint', ''),
            ep.get('access_type', ''),
            ep.get('expected_roles', ''),
            ep.get('source_file', ''),
        ], row=i, alt=alt)

        # Colour method cell
        method_cell = ws.cell(row=i, column=1)
        method = ep.get('method', '').upper()
        if method in ('FIRESTORE',):
            method_cell.fill = _fill(C['low_bg'])
        elif method in ('AUTH',):
            method_cell.fill = _fill(C['medium_bg'])
        elif method in ('HTTP',):
            method_cell.fill = _fill(C['high_bg'])
        method_cell.font = _font(bold=True, size=10)
        method_cell.alignment = Alignment(horizontal='center', vertical='top')

    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(endpoints) + 1}'
    return ws


# ─────────────────────────────────────────────────────────────
# Tab 3 — Findings
# ─────────────────────────────────────────────────────────────

def _tab_findings(wb, test_results):
    ws = wb.create_sheet('Findings')
    ws.sheet_view.showGridLines = False

    headers = [
        'ID', 'Severity', 'Endpoint', 'Method',
        'Test Category', 'Expected', 'Actual', 'Status', 'Recommendation',
    ]
    widths = [14, 12, 38, 12, 22, 30, 40, 10, 50]
    _header_row(ws, headers, row=1, widths=widths)

    for i, r in enumerate(test_results, 2):
        alt = (i % 2 == 0)
        _data_row(ws, [
            r.get('id', ''),
            r.get('severity', ''),
            r.get('endpoint', ''),
            r.get('method', ''),
            r.get('category', ''),
            r.get('expected', ''),
            r.get('actual', ''),
            r.get('status', ''),
            r.get('recommendation', ''),
        ], row=i, alt=alt)

        # Colour severity cell
        sev_cell = ws.cell(row=i, column=2)
        bg, fg = _severity_fill(r.get('severity', ''))
        if bg:
            sev_cell.fill = _fill(bg)
            sev_cell.font = _font(bold=True, color=fg, size=10)
        sev_cell.alignment = Alignment(horizontal='center', vertical='top')

        # Colour status cell
        status_cell = ws.cell(row=i, column=8)
        status = r.get('status', '')
        if status == 'PASS':
            status_cell.fill = _fill(C['pass_bg'])
            status_cell.font = _font(bold=True, color=C['pass_fg'], size=10)
        elif status == 'FAIL':
            status_cell.fill = _fill(C['fail_bg'])
            status_cell.font = _font(bold=True, color=C['fail_fg'], size=10)
        status_cell.alignment = Alignment(horizontal='center', vertical='top')

    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(test_results) + 1}'
    ws.freeze_panes = 'A2'
    return ws


# ─────────────────────────────────────────────────────────────
# Tab 4 — Remediation Status
# ─────────────────────────────────────────────────────────────

def _tab_remediation_status(wb, remediation_log):
    ws = wb.create_sheet('Remediation Status')
    ws.sheet_view.showGridLines = False

    headers = [
        'Finding ID', 'Auto Fixed', 'Fix Status', 'Files Changed',
        'Commit SHA', 'PR Number', 'Verification Status', 'Details',
    ]
    widths = [14, 12, 18, 40, 20, 12, 22, 60]
    _header_row(ws, headers, row=1, widths=widths)

    for i, r in enumerate(remediation_log, 2):
        alt = (i % 2 == 0)
        files = ', '.join(r.get('files_changed', []))
        _data_row(ws, [
            r.get('finding_id', ''),
            'Yes ✅' if r.get('auto_fixed') else 'No ❌',
            r.get('fix_status', ''),
            files,
            r.get('commit_sha', ''),
            r.get('pr_number', ''),
            r.get('verification_status', ''),
            r.get('details', '') or r.get('manual_action', ''),
        ], row=i, alt=alt)

        fixed_cell = ws.cell(row=i, column=2)
        if r.get('auto_fixed'):
            fixed_cell.fill = _fill(C['pass_bg'])
            fixed_cell.font = _font(bold=True, color=C['pass_fg'], size=10)
        else:
            fixed_cell.fill = _fill(C['medium_bg'])
            fixed_cell.font = _font(bold=True, color=C['medium_fg'], size=10)
        fixed_cell.alignment = Alignment(horizontal='center', vertical='top')

    ws.freeze_panes = 'A2'
    return ws


# ─────────────────────────────────────────────────────────────
# Tab 5 — Raw Results
# ─────────────────────────────────────────────────────────────

def _tab_raw_results(wb, test_results):
    ws = wb.create_sheet('Raw Results')
    ws.sheet_view.showGridLines = False

    headers = [
        'Timestamp', 'Test ID', 'Endpoint', 'Method',
        'Role', 'HTTP Status', 'Response Time (ms)', 'Result',
    ]
    widths = [26, 14, 40, 12, 16, 14, 20, 10]
    _header_row(ws, headers, row=1, widths=widths)

    row = 2
    for r in test_results:
        for raw in r.get('raw', []):
            alt = (row % 2 == 0)
            _data_row(ws, [
                raw.get('timestamp', ''),
                r.get('id', ''),
                raw.get('endpoint', ''),
                raw.get('method', ''),
                raw.get('role', ''),
                raw.get('http_status', ''),
                raw.get('response_time_ms', ''),
                raw.get('result', ''),
            ], row=row, alt=alt)

            result_cell = ws.cell(row=row, column=8)
            result = raw.get('result', '')
            if result == 'PASS':
                result_cell.fill = _fill(C['pass_bg'])
                result_cell.font = _font(bold=True, color=C['pass_fg'], size=10)
            elif result == 'FAIL':
                result_cell.fill = _fill(C['fail_bg'])
                result_cell.font = _font(bold=True, color=C['fail_fg'], size=10)
            result_cell.alignment = Alignment(horizontal='center', vertical='top')
            row += 1

    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{row - 1}'
    ws.freeze_panes = 'A2'
    return ws


# ─────────────────────────────────────────────────────────────
# Main entry point
# ─────────────────────────────────────────────────────────────

def generate_report(test_results, endpoints, remediation_log, output_path=None):
    """
    Generate the full 5-tab Excel evidence report.

    Args:
        test_results:    list of dicts from security_tests.run_all_tests()
        endpoints:       list of dicts from security_tests.discover_endpoints()
        remediation_log: list of dicts from auto_remediate.apply_all_fixes()
        output_path:     override output path (default: automated_test/security_test_report.xlsx)
    """
    out_path = output_path or REPORT_PATH
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    print('[Report] Building Executive Summary tab...')
    _tab_executive_summary(wb, test_results, endpoints)

    print('[Report] Building Endpoint Inventory tab...')
    _tab_endpoint_inventory(wb, endpoints)

    print('[Report] Building Findings tab...')
    _tab_findings(wb, test_results)

    print('[Report] Building Remediation Status tab...')
    _tab_remediation_status(wb, remediation_log)

    print('[Report] Building Raw Results tab...')
    _tab_raw_results(wb, test_results)

    wb.save(out_path)
    print(f'[Report] ✅ Saved: {out_path}')
    return out_path


if __name__ == '__main__':
    import sys
    sys.path.insert(0, os.path.dirname(__file__))
    from security_tests import run_all_tests, discover_endpoints
    from auto_remediate import apply_all_fixes

    print('[Report] Running tests for standalone report generation...')
    eps = discover_endpoints()
    results = run_all_tests()
    rem_log = apply_all_fixes(results)
    generate_report(results, eps, rem_log)
